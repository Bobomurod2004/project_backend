import io
from decimal import Decimal

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


def _fmt(val):
    if val is None:
        return ''
    v = Decimal(str(val))
    return f"{v:,.3f}".replace(',', ' ')


def _fmt_sum(val):
    if val is None:
        return ''
    v = Decimal(str(val))
    return f"{v:,.2f}".replace(',', ' ')


# Column definitions (key, short label for row 2 header)
USTUNLAR = [
    ('yil_boshi_hajmi',      'Hajmi'),
    ('yil_boshi_absolyut',   'Abs. spirt'),
    ('qabul_hajmi',          'Qabul'),
    ('sarflangan_hajmi',     'Sarflangan'),
    ('ishlab_hajmi',         'Hajmi'),
    ('ishlab_absolyut',      'Abs. spirt'),
    # Realizatsiya > Jami
    ('realizatsiya_hajmi',   'Hajmi'),
    ('realizatsiya_absolyut','Abs. spirt'),
    ('realizatsiya_summasi', 'Summasi'),
    # Realizatsiya > Shundan, eksport
    ('eksport_hajmi',        'Hajmi'),
    ('eksport_absolyut',     'Abs. spirt'),
    ('eksport_summasi',      'Summasi'),
    # rest
    ('yoqotish_hajmi',       "Yo'qotish"),
    ('oz_ehtiyoj_hajmi',     "O'z ehtiyoji"),
    ('oy_oxiri_hajmi',       'Hajmi'),
    ('oy_oxiri_absolyut',    'Abs. spirt'),
]

SUM_FIELDS = {'realizatsiya_summasi', 'eksport_summasi'}

# Svod (oy-ma-oy umumiy ko'rinish) ustunlari — qisqa, tushunarli sarlavhalar
SVOD_USTUNLAR = [
    ('yil_boshi_hajmi',      'Yil boshi haj.'),
    ('yil_boshi_absolyut',   'Yil boshi abs.'),
    ('qabul_hajmi',          'Qabul'),
    ('sarflangan_hajmi',     'Sarflangan'),
    ('ishlab_hajmi',         'Ishlab haj.'),
    ('ishlab_absolyut',      'Ishlab abs.'),
    ('realizatsiya_hajmi',   'Realiz. haj.'),
    ('realizatsiya_absolyut', 'Realiz. abs.'),
    ('realizatsiya_summasi', 'Realiz. sum.'),
    ('eksport_hajmi',        'Eksport haj.'),
    ('eksport_absolyut',     'Eksport abs.'),
    ('eksport_summasi',      'Eksport sum.'),
    ('yoqotish_hajmi',       "Yo'qotish"),
    ('oz_ehtiyoj_hajmi',     "O'z ehtiyoji"),
    ('oy_oxiri_hajmi',       'Oy oxiri haj.'),
    ('oy_oxiri_absolyut',    'Oy oxiri abs.'),
]

# Qoldiq ustunlari — yillik jamida yig'indi emas, chegaraviy oy qiymati olinadi
BALANS_BOSHI = {'yil_boshi_hajmi', 'yil_boshi_absolyut'}
BALANS_OXIRI = {'oy_oxiri_hajmi', 'oy_oxiri_absolyut'}


def oylik_svod_data(korxona, yil):
    """
    Korxonaning bir yildagi oylik hisobotlari bo'yicha oy-ma-oy svod.

    Har bir oy uchun barcha mahsulotlar yig'indisi olinadi; oxirida yillik
    jami. Qoldiq (yil boshi / oy oxiri) ustunlari yig'indi emas — chegaraviy
    oyning qiymati ishlatiladi.

    Qaytadi: {'yil', 'oylar': [{'oy', 'oy_nomi', <field>: float|None}, ...],
              'jami': {<field>: float|None}}
    """
    from django.db.models import Sum
    from .models import OylikHisobot

    fields = [key for key, _ in SVOD_USTUNLAR]
    hisobotlar = (
        OylikHisobot.objects
        .filter(korxona=korxona, yil=yil)
        .order_by('oy')
    )

    oylar = []
    for h in hisobotlar:
        agg = h.qatorlar.aggregate(**{f: Sum(f) for f in fields})
        row = {'oy': h.oy, 'oy_nomi': h.get_oy_display()}
        for f in fields:
            v = agg.get(f)
            row[f] = float(v) if v is not None else None
        oylar.append(row)

    jami = {}
    for f in fields:
        if f in BALANS_BOSHI:
            jami[f] = oylar[0][f] if oylar else None
        elif f in BALANS_OXIRI:
            jami[f] = oylar[-1][f] if oylar else None
        else:
            vals = [o[f] for o in oylar if o[f] is not None]
            jami[f] = sum(vals) if vals else None

    return {'yil': yil, 'oylar': oylar, 'jami': jami}


def _svod_rows_jami(rows):
    jami = {}
    for f, _ in SVOD_USTUNLAR:
        vals = [row.get(f) for row in rows if row.get(f) is not None]
        jami[f] = sum(vals) if vals else None
    return jami


def svodlar_data(yil, korxona_id=None, oy=None):
    """
    Bir yil bo'yicha svodni oylar kesimida, har oy ichida korxonalar bilan
    qaytaradi. Frontenddagi alohida Svod sahifasi shu formatdan foydalanadi.
    """
    from .models import OylikHisobot

    hisobotlar = OylikHisobot.objects.select_related('korxona').filter(yil=yil)
    if korxona_id:
        hisobotlar = hisobotlar.filter(korxona_id=korxona_id)

    tanlangan_hisobotlar = hisobotlar
    if oy:
        tanlangan_hisobotlar = tanlangan_hisobotlar.filter(oy=oy)

    hisobot_idlar = {
        (h.korxona_id, h.oy): h.id
        for h in hisobotlar
    }

    korxona_hisobotlari = {}
    for h in tanlangan_hisobotlar.order_by('korxona__nomi', 'oy'):
        if h.korxona_id not in korxona_hisobotlari:
            korxona_hisobotlari[h.korxona_id] = h

    oylar_map = {}
    korxonalar = []

    for h in korxona_hisobotlari.values():
        korxona = h.korxona
        svod = oylik_svod_data(korxona, yil)
        korxona_oylari = [
            row for row in svod['oylar']
            if not oy or row['oy'] == oy
        ]
        if not korxona_oylari:
            continue

        korxonalar.append({
            'korxona': korxona.id,
            'korxona_nomi': korxona.nomi,
            'inn': korxona.inn,
            'hisobot_id': h.id,
            'jami': svod['jami'],
        })

        for row in korxona_oylari:
            oy_key = row['oy']
            if oy_key not in oylar_map:
                oylar_map[oy_key] = {
                    'oy': oy_key,
                    'oy_nomi': row['oy_nomi'],
                    'korxonalar': [],
                }

            oylar_map[oy_key]['korxonalar'].append({
                'korxona': korxona.id,
                'korxona_nomi': korxona.nomi,
                'inn': korxona.inn,
                'hisobot_id': hisobot_idlar.get((korxona.id, oy_key)),
                **{f: row.get(f) for f, _ in SVOD_USTUNLAR},
            })

    oylar = []
    for item in sorted(oylar_map.values(), key=lambda x: x['oy']):
        item['korxonalar'].sort(key=lambda row: row['korxona_nomi'])
        item['jami'] = _svod_rows_jami(item['korxonalar'])
        oylar.append(item)

    return {
        'yil': yil,
        'oy': oy,
        'oylar': oylar,
        'jami': _svod_rows_jami([k['jami'] for k in korxonalar]),
        'korxonalar_soni': len(korxonalar),
        'hisobotlar_soni': sum(len(item['korxonalar']) for item in oylar),
    }

# Column indices (0-based, after T/p=0 and Mahsulot=1)
# yil_boshi_hajmi=2, yil_boshi_absolyut=3
# qabul=4, sarflangan=5
# ishlab_hajmi=6, ishlab_absolyut=7
# realiz_hajmi=8, realiz_abs=9, realiz_sum=10
# eksp_hajmi=11, eksp_abs=12, eksp_sum=13
# yoqotish=14, oz_ehtiyoj=15
# oy_oxiri_hajmi=16, oy_oxiri_abs=17


# ─── PDF ────────────────────────────────────────────────────────────────────

def hisobot_pdf(hisobot):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=8 * mm,
        rightMargin=8 * mm,
        topMargin=8 * mm,
        bottomMargin=8 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle('title', fontSize=10, alignment=1, spaceAfter=3)
    h_style = ParagraphStyle('hdr', fontSize=5.5, alignment=1, leading=7)
    # Mahsulot nomi — chapga tekislangan, o'raladigan (wrap) uslub
    name_style = ParagraphStyle('name', fontSize=6.5, alignment=0, leading=8)

    def P(txt):
        return Paragraph(txt, h_style)

    elements = []

    elements.append(Paragraph(
        "SPIRTLI ICHIMLIKLAR ISHLAB CHIQARISH VA REALIZATSIYA HISOBOTI",
        title_style
    ))
    elements.append(Paragraph(
        f"Korxona: {hisobot.korxona.nomi}  |  INN: {hisobot.korxona.inn}  |  "
        f"{hisobot.yil} yil {hisobot.get_oy_display()}",
        title_style
    ))
    elements.append(Spacer(1, 3 * mm))

    # 3-row header
    # Row 0: super-groups
    # Row 1: sub-groups (Realizatsiya split)
    # Row 2: column names
    row0 = [
        P('T/p'), P('Mahsulot turi'),
        P("Yil boshiga\nqoldiq"), '',
        P("Qabul /\nSarflangan"), '',
        P("Ishlab\nchiqarish"), '',
        P('Realizatsiya'), '', '', '', '', '',
        P("Yo'qotish /\nEhtiyoj"), '',
        P("Oy oxiriga\nqoldiq"), '',
    ]
    row1 = [
        '', '',
        '', '',
        '', '',
        '', '',
        P('Jami'), '', '',
        P("Shundan,\neksport"), '', '',
        '', '',
        '', '',
    ]
    row2 = [
        '', '',
        P('Hajmi'), P('Abs.\nspirt'),
        P('Qabul'), P('Sarf-\nlangan'),
        P('Hajmi'), P('Abs.\nspirt'),
        P('Hajmi'), P('Abs.\nspirt'), P('Summasi'),
        P('Hajmi'), P('Abs.\nspirt'), P('Summasi'),
        P("Yo'qotish"), P("O'z\nehtiyoji"),
        P('Hajmi'), P('Abs.\nspirt'),
    ]

    qatorlar_qs = hisobot.qatorlar.select_related('mahsulot').order_by('mahsulot__tartib_raqam')

    data = [row0, row1, row2]
    for q in qatorlar_qs:
        m = q.mahsulot
        # Mahsulot nomi — Paragraph (uzun nomlar katak ichida o'raladi)
        row = [str(m.tartib_raqam), Paragraph(m.nomi, name_style)]
        for field, _ in USTUNLAR:
            val = getattr(q, field)
            row.append(_fmt_sum(val) if field in SUM_FIELDS else _fmt(val))
        data.append(row)

    # Landshaft A4 = 297mm, chekkalar 8+8 = 16mm -> foydali ~281mm
    # 6 + 42 + 16*14.55 = 280.8mm (sahifaga sig'adi)
    col_widths = [6 * mm, 42 * mm] + [14.55 * mm] * 16

    table = Table(data, colWidths=col_widths, repeatRows=3)
    HDR_BG  = colors.HexColor('#b9c79b')
    SUB_BG  = colors.HexColor('#cdd9b4')
    COL_BG  = colors.HexColor('#e3e9d4')
    HDR_TXT = colors.HexColor('#2f3a1a')

    style_cmds = [
        # backgrounds
        ('BACKGROUND', (0, 0), (-1, 0), HDR_BG),
        ('BACKGROUND', (0, 1), (-1, 1), SUB_BG),
        ('BACKGROUND', (0, 2), (-1, 2), COL_BG),
        ('TEXTCOLOR',  (0, 0), (-1, 2), HDR_TXT),
        # alternating rows
        ('ROWBACKGROUNDS', (0, 3), (-1, -1),
         [colors.white, colors.HexColor('#f4f7ec')]),
        # fonts & align
        ('FONTSIZE', (0, 0), (-1, 2), 5.5),
        ('FONTSIZE', (0, 3), (-1, -1), 6.2),
        ('ALIGN',    (0, 0), (-1, -1), 'CENTER'),
        # Mahsulot nomi ustuni — chapga; raqamlar — o'ngga
        ('ALIGN',    (1, 3), (1, -1), 'LEFT'),
        ('ALIGN',    (2, 3), (-1, -1), 'RIGHT'),
        ('VALIGN',   (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID',     (0, 0), (-1, -1), 0.3, colors.grey),
        ('TOPPADDING',    (0, 0), (-1, -1), 2),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ('LEFTPADDING',   (0, 0), (-1, -1), 2),
        ('RIGHTPADDING',  (0, 0), (-1, -1), 2),
        # SPANS — header
        ('SPAN', (0, 0), (0, 2)),   # T/p
        ('SPAN', (1, 0), (1, 2)),   # Mahsulot turi
        ('SPAN', (2, 0), (3, 1)),   # Yil boshi
        ('SPAN', (4, 0), (5, 1)),   # Qabul/Sarflangan
        ('SPAN', (6, 0), (7, 1)),   # Ishlab chiqarish
        ('SPAN', (8, 0), (13, 0)),  # Realizatsiya (super)
        ('SPAN', (8, 1), (10, 1)),  # Jami
        ('SPAN', (11, 1), (13, 1)), # Shundan, eksport
        ('SPAN', (14, 0), (15, 1)), # Yo'qotish/Ehtiyoj
        ('SPAN', (16, 0), (17, 1)), # Oy oxiriga qoldiq
    ]

    table.setStyle(TableStyle(style_cmds))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


# ─── EXCEL ──────────────────────────────────────────────────────────────────

def hisobot_excel(hisobot):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{hisobot.yil}-{hisobot.oy:02d}"

    TOTAL_COLS = 18  # A..R

    HDR1_FILL = PatternFill('solid', fgColor='B9C79B')
    HDR2_FILL = PatternFill('solid', fgColor='CDD9B4')
    HDR3_FILL = PatternFill('solid', fgColor='E3E9D4')
    ALT_FILL  = PatternFill('solid', fgColor='F4F7EC')

    HDR_FONT  = Font(bold=True, color='2F3A1A', size=9)
    NORMAL    = Font(size=9)

    thin   = Side(style='thin',   color='AAAAAA')
    medium = Side(style='medium', color='888888')
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    MED_BORDER = Border(left=medium, right=medium, top=medium, bottom=medium)

    CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
    LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

    def hdr(row, col, value, fill, merge_to_row=None, merge_to_col=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = HDR_FONT
        c.fill = fill
        c.alignment = CENTER
        c.border = BORDER
        if merge_to_row or merge_to_col:
            ws.merge_cells(
                start_row=row, start_column=col,
                end_row=merge_to_row or row,
                end_column=merge_to_col or col,
            )

    # ── Title rows ──
    last_col = get_column_letter(TOTAL_COLS)
    ws.merge_cells(f'A1:{last_col}1')
    ws['A1'] = 'SPIRTLI ICHIMLIKLAR ISHLAB CHIQARISH VA REALIZATSIYA HISOBOTI'
    ws['A1'].font = Font(bold=True, size=12)
    ws['A1'].alignment = CENTER

    ws.merge_cells(f'A2:{last_col}2')
    ws['A2'] = (
        f"Korxona: {hisobot.korxona.nomi}  |  INN: {hisobot.korxona.inn}  |  "
        f"{hisobot.yil} yil {hisobot.get_oy_display()}"
    )
    ws['A2'].font = Font(bold=True, size=10)
    ws['A2'].alignment = CENTER

    # ── Row 3: Super-group headers ──
    ws.row_dimensions[3].height = 22
    hdr(3, 1,  'T/p',                     HDR1_FILL, merge_to_row=5)            # A3:A5
    hdr(3, 2,  'Mahsulot turi',            HDR1_FILL, merge_to_row=5)            # B3:B5
    hdr(3, 3,  'Yil boshiga qoldiq',       HDR1_FILL, merge_to_row=4, merge_to_col=4)  # C3:D4
    hdr(3, 5,  'Qabul / Sarflangan',       HDR1_FILL, merge_to_row=4, merge_to_col=6)  # E3:F4
    hdr(3, 7,  'Ishlab chiqarish',         HDR1_FILL, merge_to_row=4, merge_to_col=8)  # G3:H4
    hdr(3, 9,  'Realizatsiya',             HDR1_FILL, merge_to_col=14)           # I3:N3
    hdr(3, 15, "Yo'qotish / Ehtiyoj",     HDR1_FILL, merge_to_row=4, merge_to_col=16) # O3:P4
    hdr(3, 17, 'Oy oxiriga qoldiq',        HDR1_FILL, merge_to_row=4, merge_to_col=18) # Q3:R4

    # ── Row 4: Sub-group headers ──
    ws.row_dimensions[4].height = 18
    hdr(4, 9,  'Jami',              HDR2_FILL, merge_to_col=11)   # I4:K4
    hdr(4, 12, "Shundan, eksport",  HDR2_FILL, merge_to_col=14)   # L4:N4

    # ── Row 5: Column labels ──
    ws.row_dimensions[5].height = 30
    col_labels = [
        'Hajmi', 'Abs. spirt',       # yil boshi (C,D)
        'Qabul', 'Sarflangan',        # qabul/sarf (E,F)
        'Hajmi', 'Abs. spirt',        # ishlab (G,H)
        'Hajmi', 'Abs. spirt', 'Summasi',   # realiz jami (I,J,K)
        'Hajmi', 'Abs. spirt', 'Summasi',   # eksport (L,M,N)
        "Yo'qotish", "O'z ehtiyoji",  # (O,P)
        'Hajmi', 'Abs. spirt',        # oy oxiri (Q,R)
    ]
    for i, label in enumerate(col_labels, start=3):
        hdr(5, i, label, HDR3_FILL)

    # ── Data rows ──
    qatorlar_qs = hisobot.qatorlar.select_related('mahsulot').order_by('mahsulot__tartib_raqam')

    for row_idx, q in enumerate(qatorlar_qs, start=6):
        m = q.mahsulot
        fill = ALT_FILL if row_idx % 2 == 0 else None

        def wcell(col, value, num_fmt=None):
            c = ws.cell(row=row_idx, column=col, value=value)
            c.font = NORMAL
            c.border = BORDER
            c.alignment = CENTER if num_fmt else LEFT
            if fill:
                c.fill = fill
            if num_fmt and value is not None:
                c.number_format = num_fmt
            return c

        wcell(1, m.tartib_raqam)
        wcell(2, m.nomi)

        for col_offset, (field, _) in enumerate(USTUNLAR, start=3):
            val = getattr(q, field)
            if val is not None:
                val = float(val)
            fmt = '#,##0.00' if field in SUM_FIELDS else '#,##0.000'
            wcell(col_offset, val, num_fmt=fmt if val is not None else None)

    # ── Column widths ──
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 28
    for col_idx in range(3, TOTAL_COLS + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()
